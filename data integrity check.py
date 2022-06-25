import openpyxl
import os

file_path = os.getcwd()
wb_data = openpyxl.load_workbook('偿二代二期汇总表粗并.xlsx')
ws_data = wb_data.worksheets[0]
dict_asset_type = {}
wb_problem = openpyxl.Workbook()
ws_problem = wb_problem.worksheets[0]
ls_problem_index = []


def initialize_problem_sheet():
    for col in range(ws_data.max_column):
        ws_problem.cell(row=1, column=col + 1).value = ws_data.cell(row=1, column=col + 1).value
    ws_problem.cell(row=1, column=ws_data.max_column+1).value = "识别"
    ws_problem.cell(row=1, column=ws_data.max_column+2).value = "问题"


def show_problem(problem_id_list, problem):
    for row_problem in range(ws_problem.max_row - 1):
        if ws_problem.cell(row=row_problem + 2, column=ws_data.max_column+1).value in problem_id_list:
            if ws_problem.cell(row=row_problem + 2, column=ws_data.max_column+2).value is not None:
                ws_problem.cell(row=row_problem + 2, column=ws_data.max_column+2).value = ws_problem.cell(
                    row_problem + 2, column=ws_data.max_column).value + "\n" + problem
            else:
                ws_problem.cell(row=row_problem + 2, column=ws_data.max_column+2).value = problem


def paste_data_with_problem(problem_index_list):
    row_problem = 2
    for row_num in problem_index_list:
        for col in range(ws_data.max_column):
            if ws_data.cell(row=row_num, column=col + 1).value is not None:
                ws_problem.cell(row=row_problem, column=col + 1).value = ws_data.cell(row=row_num, column=col + 1).value
        ws_problem.cell(row=row_problem, column=ws_data.max_column+1).value = row_num
        row_problem += 1


def check_none(col):
    ls_result = []
    for row_check_none in range(2, ws_data.max_row + 1):
        if ws_data.cell(row=row_check_none, column=col).value is None:
            ls_result.append(row_check_none)
    return ls_result


def check_value_compliance(col, ls_rule):
    ls_result = []
    for j in range(ws_data.max_row - 1):
        cell_value = ws_data.cell(row=j + 2, column=col).value
        bull_data = 0
        for item in ls_rule:
            if cell_value == item:
                bull_data = 1
        if bull_data == 0:
            ls_result.append(j + 2)
    return ls_result


def check_none_by_type(asset_type_check):
    dict_result_type_exist = {}
    dict_result_type_null = {}
    for row_asset in range(2, ws_data.max_row+1):
        ls_rows_to_check = []
        if ws_data.cell(row=row_asset, column=4).value == asset_type_check:
            ls_rows_to_check.append(row_asset)
    ls_col_to_check_exist = dict_asset_type_existence[asset_type_check]
    ls_col_to_check_null = dict_asset_type_null[asset_type_check]
    for row_asset_to_check in ls_rows_to_check:
        ls_result_exist = []
        ls_result_null = []
        bull_exist = 0
        bull_null = 0
        for col_to_check in ls_col_to_check_exist:
            if ws_data.cell(row=row_asset_to_check, column=col_to_check).value is None:
                ls_result_exist.append(col_to_check)
                bull_exist = 1
        if bull_exist == 1:
            dict_result_type_exist[row_asset_to_check] = ls_result_exist
        for col_to_check in ls_col_to_check_null:
            if ws_data.cell(row=row_asset_to_check, column=col_to_check).value is not None:
                ls_result_null.append(row_asset_to_check)
                bull_null = 1
        if bull_null == 1:
            dict_result_type_null[row_asset_to_check] = ls_result_null
        dict_result_exist[asset_type_check] = dict_result_type_exist
        dict_result_null[asset_type_check] = dict_result_type_null


# 大类与类型的取值规范字典
for i in range(wb_data.worksheets[1].max_row - 2):
    asset_type_ = wb_data.worksheets[1].cell(row=i + 3, column=2).value
    asset_type_general_ = wb_data.worksheets[1].cell(row=i + 3, column=1).value
    dict_asset_type[asset_type_] = asset_type_general_
ls_account_ = ["万能", "传统"]

ls_type = check_value_compliance(4, dict_asset_type.keys())
ls_type_general = check_value_compliance(3, dict_asset_type.values())
ls_account = check_value_compliance(4, ls_account_)

# 检查大类与类型是否匹配
ls_match_type_raw = []
ls_match_type = []
for i in range(2, ws_data.max_row + 1):
    if i not in ls_type:
        ls_match_type_raw.append(i)
for row in ls_match_type_raw:
    key = ws_data.cell(row=row, column=4).value
    if dict_asset_type[key] != ws_data.cell(row=row, column=3).value:
        ls_match_type.append(row)

# 检查1 2 3 4 8 9 12 66列是否为空
ls_not_null_1 = []
ls_not_null_2 = []
ls_not_null_3 = []
ls_not_null_4 = []
ls_not_null_8 = []
ls_not_null_9 = []
ls_not_null_12 = []
ls_not_null_66 = []
ls_not_null = [ls_not_null_1, ls_not_null_2, ls_not_null_3, ls_not_null_4, ls_not_null_8, ls_not_null_9, ls_not_null_12,
               ls_not_null_66]
ls_not_null_num = [1, 2, 3, 4, 8, 9, 12, 66]
ls_not_null_name = ["资产简称", "资产全称", "资产大类", "资产类型", "购买成本", "认可价值", "账户", "资产五大类分类"]
ls_count = 0
for ls_result in ls_not_null:
    ls_result = check_none(ls_not_null_num[ls_count])
    ls_count += 1

# 提取对应资产的空列与非空列
dict_asset_type_existence = {}
dict_asset_type_null = {}
ws_data_existence = wb_data.worksheets[4]
for row_existence in range(2, ws_data_existence.max_row+1):
    asset_type = ws_data_existence.cell(row=row_existence, column=1).value
    list_col_exist = []
    list_col_null = []
    for col_existence in range(2, ws_data_existence.max_column+1):
        if ws_data_existence.cell(row=row_existence, column=col_existence).value == 1:
            list_col_exist.append(col_existence-1)
        else:
            if ws_data_existence.cell(row=row_existence, column=col_existence).value == 0:
                list_col_null.append(col_existence-1)
    dict_asset_type_existence[asset_type] = list_col_exist
    dict_asset_type_null[asset_type] = list_col_null

dict_result_exist = {}
dict_result_null = {}

for type_of_asset in dict_asset_type.keys():
    check_none_by_type(type_of_asset)


# ls_problem_index = []
ls_problem_index.extend(ls_type)
ls_problem_index.extend(ls_type_general)
ls_problem_index.extend(ls_match_type)
for ls in ls_not_null:
    ls_problem_index.extend(ls)
for values in dict_result_exist.values():
    ls_problem_index.extend(values.keys())

ls_problem_index = list(set(ls_problem_index))
ls_problem_index.sort()
# initialize_problem_sheet()
# paste_data_with_problem(ls_problem_index)
#
# show_problem(ls_type, "资产类型不符合规定")
# show_problem(ls_type_general, "资产大类不符合规定")
# count_ls_not_null = 0
# for ls in ls_not_null:
#     show_problem(ls, ls_not_null_name[count_ls_not_null] + "不应为空值")
#     count_ls_not_null += 1
#
asset_type_check = "沪深主板股"
dict_result_type_exist = {}
dict_result_type_null = {}
for row_asset in range(2, ws_data.max_row + 1):
    ls_rows_to_check = []
    if ws_data.cell(row=row_asset, column=4).value == asset_type_check:
        ls_rows_to_check.append(row_asset)
ls_col_to_check_exist = dict_asset_type_existence[asset_type_check]
ls_col_to_check_null = dict_asset_type_null[asset_type_check]
for row_asset_to_check in ls_rows_to_check:
    ls_result_exist = []
    ls_result_null = []
    bull_exist = 0
    bull_null = 0
    for col_to_check in ls_col_to_check_exist:
        if ws_data.cell(row=row_asset_to_check, column=col_to_check).value is None:
            ls_result_exist.append(col_to_check)
            bull_exist = 1
    print("行")
    if bull_exist == 1:
        dict_result_type_exist[row_asset_to_check] = ls_result_exist
        print(row_asset_to_check)
        print(ls_result_exist)
    for col_to_check in ls_col_to_check_null:
        if ws_data.cell(row=row_asset_to_check, column=col_to_check).value is not None:
            ls_result_null.append(row_asset_to_check)
            bull_null = 1
    print(ls_result_exist)
    if bull_null == 1:
        dict_result_type_null[row_asset_to_check] = ls_result_null
        print(row_asset_to_check)
        print(ls_result_exist)
    dict_result_exist[asset_type_check] = dict_result_type_exist
    dict_result_null[asset_type_check] = dict_result_type_null

