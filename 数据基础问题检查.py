import datetime
import time
from datetime import datetime
import openpyxl
import os


file_name = input('请输入数据文件名：')
file_path = os.getcwd()
if file_name[-5:] == '.xlsx':
    file_name = file_name[-len(file_name):-5]
wb_data = openpyxl.load_workbook(file_path + '/'+file_name + '.xlsx')
ws_data = wb_data.worksheets[0]
wb_data_rule = openpyxl.load_workbook(file_path + '/数据规则.xlsx')
ws_rule_null = wb_data_rule['字段空与非空']
ws_rule_type = wb_data_rule['数据类型']
ws_valid_value = wb_data_rule['字段取值']
dict_col_name = {}
dict_name_col = {}
dict_asset_type = {}
ls_problem_index = []
dict_check_null_by_type = {}
dict_check_not_null_by_type = {}
dict_col_type = {}
dict_valid_value = {}
dict_lower_limit = {}
dict_upper_limit = {}


def initialize_problem_sheet():
    ws_data.cell(row=1, column=ws_data.max_column + 1).value = "问题"


def mark_the_problem(row_num, ls_problem):
    string_problem = str()
    problem_counts = 1
    for problems in ls_problem:
        string_problem += str(problem_counts) + "." + problems + '\n'
        problem_counts += 1
    ws_data.cell(row=row_num, column=ws_data.max_column).value = string_problem


def check_value_not_null(row_num, col_num):
    if ws_data.cell(row=row_num, column=col_num).value is None:
        ls_problem.append(dict_col_name[col_num] + '为空')
        global bull_problem
        bull_problem = 1


def check_value_valid(row_num, col_num, ls_valid_value):
    cell_value = ws_data.cell(row=row_num, column=col_num).value
    bull_valid_value = 0
    if cell_value in ls_valid_value:
        return 0
    else:
        ls_problem.append(dict_col_name[col_num] + '非法取值')
        return 1


def check_value_null(row_num, col_num):
    if ws_data.cell(row=row_num, column=col_num).value is not None:
        return '第' + str(row_num) + '行' + [col_num] + '应为空'


def check_value_type(row_num, col_num, type_):
    global bull_problem
    if type_ == 'int':
        if isinstance(ws_data.cell(row=row_num, column=col_num).value, int) \
                or isinstance(ws_data.cell(row=row_num,column=col_num).value, float) \
                or ws_data.cell(row=row_num, column=col_num).value in ls_int_exception:

            pass
        else:
            ls_problem.append(dict_col_name[col_num] + '数据类型错误')

            bull_problem = 1
    else:
        if isinstance(ws_data.cell(row=row_num, column=col_num).value, eval(type_)):

            pass
        elif dict_col_name[col_num] == "产品代码" or dict_col_name[col_num] == '表层资产产品代码':
            pass
        else:
            ls_problem.append(dict_col_name[col_num] + '数据类型错误')
            bull_problem = 1


def check_number_relation_between_columns(row_num, col_smaller, col_bigger):
    if ws_data.cell(row=row_num, column=col_smaller).value > ws_data.cell(row=row_num, column=col_bigger).value:
        ls_problem.append(dict_col_name[col_smaller] + '不应大于' + dict_col_name[col_bigger])
    else:
        pass


def check_lower_limit(row_num, col_num, limit=0):
    if ws_data.cell(row=row_num, column=col_num).value < limit:
        ls_problem.append(dict_col_name[col_num] + '不应小于' + str(limit))
    else:
        pass


def check_upper_limit(row_num, col_num, limit):
    if ws_data.cell(row=row_num, column=col_num).value > limit:
        ls_problem.append(dict_col_name[col_num] + '不应大于' + str(limit))
    else:
        pass


def bull_not_str_and_null(col_num):
    value = ws_data.cell(row=row_count, column=col_num).value
    return value is not None and (isinstance(value, int) or isinstance(value, float))


##########################################################################################

time_start = time.time()

for col_count in range(66):  # range(ws_data.max_column)
    dict_col_name[col_count + 1] = ws_data.cell(row=1, column=col_count + 1).value
    dict_name_col[ws_data.cell(row=1, column=col_count + 1).value] = col_count + 1

for row_num in range(2, ws_rule_null.max_row + 1):
    type_name = ws_rule_null.cell(row=row_num, column=1).value
    ls_not_null = []
    ls_null = []
    for col_num in range(2, ws_rule_null.max_row - 1):
        if ws_rule_null.cell(row=row_num, column=col_num).value == 1:
            ls_not_null.append(col_num - 1)
        if ws_rule_null.cell(row=row_num, column=col_num).value == 0:
            ls_null.append(col_num - 1)
    dict_check_null_by_type[type_name] = ls_null
    dict_check_not_null_by_type[type_name] = ls_not_null

for row_num_ in range(1, ws_rule_type.max_row + 1):
    dict_col_type[ws_rule_type.cell(row=row_num_, column=1).value] = ws_rule_type.cell(row=row_num_, column=3).value
    if ws_rule_type.cell(row=row_num_, column=4).value is not None:
        dict_lower_limit[ws_rule_type.cell(row=row_num_, column=1).value] = ws_rule_type.cell(row=row_num_,
                                                                                              column=4).value
    if ws_rule_type.cell(row=row_num_, column=5).value is not None:
        dict_upper_limit[ws_rule_type.cell(row=row_num_, column=1).value] = ws_rule_type.cell(row=row_num_,
                                                                                              column=5).value

for col_num in range(1, ws_valid_value.max_column + 1):
    col_name = ws_valid_value.cell(row=1, column=col_num).value
    ls_valid_value = []
    for row_num__ in range(2, ws_valid_value.max_row + 1):
        if ws_valid_value.cell(row=row_num__, column=col_num).value is not None:
            ls_valid_value.append(ws_valid_value.cell(row=row_num__, column=col_num).value)
    dict_valid_value[col_name] = ls_valid_value

ls_int_exception = ['无', '无评级', '-']

initialize_problem_sheet()

#################################################################################################################################################
for row_count in range(2, ws_data.max_row + 1):
    ls_problem = []
    bull_problem = 0
    bull_asset_type = 0
    if ws_data.cell(row=row_count, column=1).value is not None:
        check_value_not_null(row_count, dict_name_col['资产类型'])
        bull_asset_type = check_value_valid(row_count, dict_name_col['资产类型'], dict_valid_value['资产类型'])
        if bull_asset_type == 0:  # 资产类型取值无误
            asset_type = ws_data.cell(row=row_count, column=4).value
            for cols in dict_col_name.keys():
                if cols in dict_check_not_null_by_type[asset_type]:
                    check_value_not_null(row_count, cols)
                if asset_type == "优先股" or asset_type == "无固定期限资产债券":
                    if dict_name_col['发行机构类型'] == '银行':
                        map(lambda col: check_value_not_null(row_count, col), [30,31,32,33])
                    elif dict_name_col['发行机构类型'] == '保险':
                        map(lambda col: check_value_not_null(row_count, col), [34,35])
                    else:
                        pass

                if ws_data.cell(row=row_count, column=cols).value is not None:
                    if bull_problem == 0:  # 对应列非空
                        check_value_type(row_count, cols, dict_col_type[cols])
                        if bull_problem == 0:  # 对应列数据类型正确
                            if dict_col_name[cols] in dict_valid_value.keys():
                                check_value_valid(row_count, cols, dict_valid_value[dict_col_name[cols]])
            if bull_not_str_and_null(dict_name_col['剩余年限']) and bull_not_str_and_null(dict_name_col['修正久期']):
                if ws_data.cell(row=row_count, column=dict_name_col['剩余年限']).value >= 1:
                    check_number_relation_between_columns(row_count, dict_name_col['修正久期'], dict_name_col['剩余年限'])
            if bull_not_str_and_null(dict_name_col['发行银行 核心一级 资本充足率']) and bull_not_str_and_null(
                    dict_name_col['发行银行一级资本充足率']):
                check_number_relation_between_columns(row_count, dict_name_col['发行银行 核心一级 资本充足率'],
                                                      dict_name_col['发行银行一级资本充足率'])
            if bull_not_str_and_null(dict_name_col['发行银行一级资本充足率']) and bull_not_str_and_null(
                    dict_name_col['发行银行资本充足率']):
                check_number_relation_between_columns(row_count, dict_name_col['发行银行一级资本充足率'], dict_name_col['发行银行资本充足率'])
            if bull_not_str_and_null(dict_name_col['发行保险公司核心偿付能力充足率']) and bull_not_str_and_null(
                    dict_name_col['发行保险公司综合偿付能力充足率']):
                check_number_relation_between_columns(row_count, dict_name_col['发行保险公司核心偿付能力充足率'],
                                                      dict_name_col['发行保险公司综合偿付能力充足率'])
            if bull_not_str_and_null(dict_name_col['认可价值']) and bull_not_str_and_null(dict_name_col['减值前账面价值']):
                check_number_relation_between_columns(row_count, dict_name_col['认可价值'], dict_name_col['减值前账面价值'])
            for keys in dict_lower_limit.keys():
                if bull_not_str_and_null(keys):
                    check_lower_limit(row_count, keys, dict_lower_limit[keys])
            for keys in dict_upper_limit.keys():
                if bull_not_str_and_null(keys):
                    check_upper_limit(row_count, keys, dict_upper_limit[keys])
    mark_the_problem(row_count, ls_problem)
    print(row_count)
input('请确认已关闭problem_sheet.xlsx 任意键继续')
wb_data.save(file_path + '/problem_sheet.xlsx')

time_end = time.time()
print('用时{}秒'.format(time_end - time_start))
input('已完成 任意键退出')
