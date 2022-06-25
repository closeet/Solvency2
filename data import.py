import openpyxl
from 字段处理 import ws_cell
from 字段处理 import max_row
from parameters import dict_current_deposit_col_name_self, dict_current_deposit_col_value, ls_current_deposit, \
    dict_bank_counter_party, ls_deposit_col_name, dict_deposit_col_name_self


def complete_data(data: dict, col: list):
    data_complete = {}
    for key in col:
        if key in col:
            data_complete[key] = data[key]
        else:
            data_complete[key] = None


def bank_counter_party(bank_name_raw: str, dict_bank_name: dict):
    counter_party = str()
    if bank_name_raw.find('行')+1:
        counter_party = bank_name_raw[:bank_name_raw.find('行')+1]
    elif bank_name_raw.find('社保账户')+1:
        counter_party = bank_name_raw[:bank_name_raw.find('社保账户')+4]
    elif bank_name_raw.find('信用合作联社')+1:
        counter_party = bank_name_raw[:bank_name_raw.find('信用合作联社')+6]
    else:
        print("识别资产名失败")
    return dict_bank_name[counter_party]


def account_categorize(account_name: str):
    if account_name is not None and account_name.find('万能')+1:
        return '万能'
    else:
        return '传统'


def import_deposit(ws, dict_col_name_corr: dict, ls_colname: list):
    dict_data_raw = {}
    dict_data = {}
    ls_col_name_raw = [ws_cell(ws, 1, col_num) for col_num in range(1, ws.max_column+1)]
    for row_num in range(2, max_row(ws)+1):
        dict_row = {}
        ls_row = [ws_cell(ws, row_num, col_num) for col_num in range(1, ws.max_column+1)]
        dict_row_raw = dict(zip(ls_col_name_raw, ls_row))
        dict_data_raw[row_num-1] = dict_row_raw
        for col in ls_colname:
            if col in dict_col_name_corr:
                dict_row[col] = dict_row_raw[dict_col_name_corr[col]]
            elif col == '资产简称' or col == '资产全称':
                dict_row[col] = dict_row_raw['开户行'] + '-' + dict_row_raw['帐号'][-4:] + '-' + \
                                str(dict_row_raw['起息日'].year)[-2:] + \
                                str(dict_row_raw['起息日'].month).rjust(2, '0') + \
                                str(dict_row_raw['起息日'].day).rjust(2, '0')
            elif col == '资产大类':
                dict_row[col] = '银行存款'
            elif col == '账户':
                dict_row[col] = account_categorize(dict_row_raw['账户'])
            elif col == '交易对手':
                dict_row[col] = bank_counter_party(dict_row_raw['开户行'], dict_bank_counter_party)
            elif col == '银行资本充足率':
                dict_row[col] = dict_row_raw['资本充足率（%）']/100
            elif col == '资产五大类分类':
                dict_row[col] = '固定收益类资产'
            else:
                dict_row[col] = None
        print(dict_row)
        dict_data[row_num-1] = dict_row
    return [dict_data, dict_data_raw]


def import_current_deposit(ws, dict_col_name_corr: dict, dict_col_value: dict, ls_colname: list):
    dict_data_raw = {}
    dict_data = {}
    ls_col_name_raw = [ws_cell(ws, 3, col_num) for col_num in range(1, ws.max_column+1)]
    for row_num in range(4, max_row(ws)+1):
        dict_row = {}
        ls_row = [ws_cell(ws, row_num, col_num) for col_num in range(1, ws.max_column+1)]
        dict_row_raw = dict(zip(ls_col_name_raw, ls_row))
        dict_data_raw[row_num-3] = dict_row_raw
        for col in ls_colname:
            if col in dict_col_name_corr:
                dict_row[col] = dict_row_raw[dict_col_name_corr[col]]
            elif col in dict_col_value:
                dict_row[col] = dict_col_value[col]
            elif col == '账户':
                dict_row[col] = account_categorize(dict_row_raw['机构名称'])
            elif col == '交易对手':
                dict_row[col] = bank_counter_party(dict_row_raw['帐户信息'], dict_bank_counter_party)
            else:
                dict_row[col] = None
        print(dict_row)
        dict_data[row_num-1] = dict_row
    return [dict_data, dict_data_raw]


wb_deposit_data = openpyxl.load_workbook('资产端数据/定期存款-2022年5月.xlsx', data_only=True)
wb_current_deposit_date = openpyxl.load_workbook('资产端数据/活期存款科目余额表.xlsx')
ws_deposit_data = wb_deposit_data['定期存款和存出资本保证金应收利息明细']
ws_current_deposit_data = wb_current_deposit_date.worksheets[0]
import_deposit(ws_deposit_data, dict_deposit_col_name_self, ls_deposit_col_name)
import_current_deposit(ws_current_deposit_data, dict_current_deposit_col_name_self, dict_current_deposit_col_value,
                       ls_current_deposit)
