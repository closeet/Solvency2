import openpyxl
import os

wb_model = openpyxl.load_workbook('20220530资产端最低资本-0601.xlsx')
wb_data = openpyxl.load_workbook('偿二代530投资财务+资管穿透（组合根据精算反馈再修改版本）.xlsx')
wb_data_non_invest = openpyxl.load_workbook(os.getcwd()+'/资产端数据/数据模板与说明（非投资资产）.xlsx')
wb_liquidity = openpyxl.load_workbook(os.getcwd()+'/资产端数据/活期存款科目余额表.xlsx', data_only=True)
wb_deposit = openpyxl.load_workbook(os.getcwd()+'/资产端数据/定期存款-2022年5月.xlsx', data_only=True)
wb_formula = openpyxl.load_workbook('公式索引.xlsx')
dict_formula = {}
ws_formula = wb_formula.worksheets[0]
ws_model = wb_model['资产汇总']
ws_data = wb_data['投资财务+资管合并']
ws_data_non_invest  =wb_data_non_invest['填写模板']
ws_data_deposit = wb_deposit['定期存款和存出资本保证金应收利息明细']
ws_data_current_deposit = wb_liquidity.worksheets[0]

dict_col_col = {1:3, 3:6, 4:7, 5:8, 6:17, 8:18, 9:19, 10:20, 13:2, 18:5, 26:9, 51:10, 52:11, 53:12, 41:13, 30:14, 49:15, 50:15, 10:22, 12:23, 66:26}
dict_col_deposit = {2:3, 19:6, 18:7, 5:15, 6:16, 11:19, 12:20, 4:17}
dict_col_current_deposit = {7:3, 12:6, 11:7, 10:17, 8:19}


def fill_formula():
    for rows_formula in range(1, ws_formula.max_row+1):
        dict_formula[ws_formula.cell(row=rows_formula, column=1).value] = ws_formula.cell(row=rows_formula, column=3).value
    for rows_model in range(6, ws_model.max_row+1):
        if ws_model.cell(row=rows_model, column=7).value in dict_formula.keys():
            if dict_formula[ws_model.cell(row=rows_model, column=7).value] == 1:
                ws_model.cell(row=rows_model, column=27).value = \
                    "=VLOOKUP(G"+str(rows_model)+",资产因子!$M$4:$N$95,2,0)"
            elif dict_formula[ws_model.cell(row=rows_model, column=7).value] == 2:
                ws_model.cell(row=rows_model, column=27).value = \
                    '=(O'+str(rows_model)+'="股份制商业银行") * ((P'+str(rows_model) + ' >= 0.12) * 0.03 + (P'+str(rows_model)+' < 0.12) * 0.05) + (O'+str(rows_model)+'="城市商业银行") * ((P'+str(rows_model)+' < 0.12) * 0.15 + AND((P'+str(rows_model)+' >= 0.12), P'+str(rows_model)+' < 0.125) * 0.1 + AND((P'+str(rows_model)+' >= 0.125), (P'+str(rows_model)+' < 0.135)) * 0.08 + (P'+str(rows_model)+' > 0.135) * 0.04) + (O'+str(rows_model)+'="国有大型商业银行") * 0.005 + (O'+str(rows_model)+'="农村商业银行") * ((P'+str(rows_model)+' < 0.125) * 0.18 + AND((P'+str(rows_model)+' >= 0.125), (P'+str(rows_model)+' < 0.135)) * 0.15 + AND((P'+str(rows_model)+' >= 0.135), (P'+str(rows_model)+' < 0.145)) * 0.1 + (P'+str(rows_model)+' >= 0.145) * 0.08)'
            elif dict_formula[ws_model.cell(row=rows_model, column=7).value] == 3:
                ws_model.cell(row=rows_model, column=27).value = \
                    '=AND($L'+str(rows_model)+'>0,$L'+str(rows_model)+'<=5)*$L'+str(rows_model)+'*(-0.0012*$L'+str(rows_model)+'+0.012)+($L'+str(rows_model)+'>5)*($L'+str(rows_model)+'*0.001+0.025)'
            elif dict_formula[ws_model.cell(row=rows_model, column=7).value] == 4:
                ws_model.cell(row=rows_model, column=27).value = \
                    '=($K'+str(rows_model)+'="AAA")*(AND($L'+str(rows_model)+'<=5,$L'+str(rows_model)+'>0)*($L'+str(rows_model)+'*(0.0006*$L'+str(rows_model)+'+0.012))+($L'+str(rows_model)+'>5)*$L'+str(rows_model)+'*0.015)+($K'+str(rows_model)+'="AA+")*(AND($L'+str(rows_model)+'<=5,$L'+str(rows_model)+'>0)*($L'+str(rows_model)+'*(0.0007*$L'+str(rows_model)+'+0.0165))+($L'+str(rows_model)+'>5)*$L'+str(rows_model)+'*0.02)+($K'+str(rows_model)+'="AA")*(AND($L'+str(rows_model)+'<=5,$L'+str(rows_model)+'>0)*($L'+str(rows_model)+'*(0.0009*$L'+str(rows_model)+'+0.025))+($L'+str(rows_model)+'>5)*$L'+str(rows_model)+'*0.0295)+($K'+str(rows_model)+'="AA-")*(AND($L'+str(rows_model)+'<=5,$L'+str(rows_model)+'>0)*($L'+str(rows_model)+'*(0.001*$L'+str(rows_model)+'+0.033))+($L'+str(rows_model)+'>5)*$L'+str(rows_model)+'*0.038)+($K'+str(rows_model)+'="A+")*(AND($L'+str(rows_model)+'<=5,$L'+str(rows_model)+'>0)*($L'+str(rows_model)+'*(0.002*$L'+str(rows_model)+'+0.04))+($L'+str(rows_model)+'>5)*$L'+str(rows_model)+'*0.05)+($K'+str(rows_model)+'="A")*(AND($L'+str(rows_model)+'<=5,$L'+str(rows_model)+'>0)*($L'+str(rows_model)+'*(0.003*$L'+str(rows_model)+'+0.045))+($L'+str(rows_model)+'>5)*$L'+str(rows_model)+'*0.06)+($K'+str(rows_model)+'="A-")*(AND($L'+str(rows_model)+'<=5,$L'+str(rows_model)+'>0)*($L'+str(rows_model)+'*(0.004*$L'+str(rows_model)+'+0.05))+($L'+str(rows_model)+'>5)*$L'+str(rows_model)+'*0.07)+($K'+str(rows_model)+'="BBB+")*(AND($L'+str(rows_model)+'<=5,$L'+str(rows_model)+'>0)*($L'+str(rows_model)+'*(0.005*$L'+str(rows_model)+'+0.05))+($L'+str(rows_model)+'>5)*$L'+str(rows_model)+'*0.075)+($K'+str(rows_model)+'="BBB")*(AND($L'+str(rows_model)+'<=5,$L'+str(rows_model)+'>0)*($L'+str(rows_model)+'*(0.01*$L'+str(rows_model)+'+0.05))+($L'+str(rows_model)+'>5)*$L'+str(rows_model)+'*0.1)'
        # print(rows_model)


def fill_data_asset():
    for row_counts in range(1,ws_data.max_row+1):
        for col_num in dict_col_col.keys():
            ws_model.cell(row=row_counts+5, column=dict_col_col[col_num]).value = ws_data.cell(row=row_counts+1, column=col_num).value


def fill_data_non_invest_asset():
    for rows in range(6, ws_model.max_row+1):
        if ws_model.cell(row=rows, column=3).value is None:
            endline = rows
            break
    print(endline)
    # print(ws_data_non_invest.max_row+1)
    for row_counts in range(3, ws_data_non_invest.max_row+1):
        for col_num in dict_col_col.keys():
            # print(endline-2+row_counts)
            # print(ws_data_non_invest.cell(row=row_counts, column=col_num).value)
            ws_model.cell(row=endline-3+row_counts, column=dict_col_col[col_num]).value = ws_data_non_invest.cell(row=row_counts, column=col_num).value

def fill_data_deposit():
    for rows in range(6, ws_model.max_row + 1):
        if ws_model.cell(row=rows, column=3).value is None:
            endline = rows
            break
    for row_counts in range(2, ws_data_deposit.max_row + 1):
        for col_num in dict_col_deposit.keys():
            print(endline-2+row_counts)
            print(ws_data_deposit.cell(row=row_counts, column=col_num).value)
            ws_model.cell(row=endline - 2 + row_counts,
                          column=dict_col_deposit[col_num]).value = ws_data_deposit.cell(row=row_counts,
                                                                                            column=col_num).value

def fill_data_current_deposit():
    for rows in range(6, ws_model.max_row + 1):
        if ws_model.cell(row=rows, column=3).value is None:
            endline = rows
            break
    for row_counts in range(4, ws_data_current_deposit.max_row + 1):
        for col_num in dict_col_current_deposit.keys():
            print(endline-4+row_counts)
            print(ws_data_deposit.cell(row=row_counts, column=col_num).value)
            ws_model.cell(row=endline - 4 + row_counts,
                          column=dict_col_current_deposit[col_num]).value = ws_data_current_deposit.cell(row=row_counts,
                                                                                            column=col_num).value








fill_data_asset()
fill_data_non_invest_asset()
fill_data_deposit()
fill_data_current_deposit()
fill_formula()
wb_model.save('自动填充公式.xlsx')



